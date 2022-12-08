import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import json
import time
from requests.exceptions import ProxyError
from multiprocessing import Pool, Value


class Gallery:
    """
    A class to get the details of a single gallery.

    Attributes:
        gallery_id (int) : Got from gallery list.
        meta (dict) : Information in json object `pageMetaInfo`.
        detail (dict) : Handled information.
    """

    def __init__(self, gallery_id):
        self.gallery_id = gallery_id

    def get_detail(self):
        """
        Get gallery detail by sending a request.
        """

        # Request URL
        gallery_url = f"https://www.artbasel.com/catalog/gallery/{self.gallery_id}"
        
        # -------- Send the request again if failed, try `max_try` times at most --------
        success = False
        max_try = 10
        while not success and max_try > 0:
            max_try -= 1

            try:
                response = requests.get(gallery_url, headers = headers)
                # resonse status is not 200
                if response.status_code != 200:
                    continue
                # respond successfully
                success = True
            
            # Encounter bad connection
            except ProxyError:
                continue
        
        if not success:
            print(f"bad connection: https://www.artbasel.com/catalog/gallery/{self.gallery_id}")

        # ------------ Extract information from the response ------------
        # Parse HTML
        soup = BeautifulSoup(response.text, "html.parser")
        # Find the useful segment and parse it into json object
        self.meta = json.loads(soup.find(id = "__NEXT_DATA__").text)["props"]["pageProps"]["pageMetaInfo"]
        # Extract and handle wanted information
        self.detail = {
            # Use dict.get method to set a default value, in case the key doesn't exist
            "display_name": self.meta.get("displayName", "NA"),
            "description": self.meta.get("description", "NA"),
            "show_names": self.handle_exibitions(),
            "email_address": self.meta.get("emailAddress", "NA"),
            "directors": self.handle_directors(),
            "addresses": self.handle_addresses(),
            "website": self.meta.get("website", "NA")
        }

    def handle_exibitions(self):
        """
        Extract and format exibitions information
        """
        # account_id = self.meta.get("accountId", "NA"),
        # exibition_url = f"https://www.artbasel.com/msvc/v1/exhibitorprofile/items/ba168459-e063-e211-b62e-00155d35011a/exhibitions?limit=3"
        
        return ", ".join(self.meta.get("showNames", []))

    def handle_directors(self):
        """
        Extract and format directors information
        """
        directors_json = self.meta.get("directors", {})
        directors = []
        for director_json in directors_json:
            name = []
            props = ["firstName", "lastName"]
            for prop in props:
                if prop in director_json:
                    name.append(director_json[prop])
            name = " ".join(name)

            director_props = [name]
            props = ["email", "fax", "mobile", "telephone"]
            for prop in props:
                if prop in director_json:
                    director_props.append(f"{prop}: {director_json[prop]}")
            directors.append('\n'.join(director_props))
        return '\n\n'.join(directors)

    def handle_addresses(self):
        """
        Extract and format addresses information
        """
        props = ["title", "address", "zipCode", "country", "phoneNumber", "emailAddress", "openingTimes"]
        addresses = []
        for address in self.meta.get("addresses", {}):
            address_props = []
            for prop in props:
                if prop in address:
                    address_props.append(address[prop])
            addresses.append('\n'.join(address_props))
        return "\n\n".join(addresses)

    def write(self, row, sheet):
        """
        Write extracted details in an excel sheet

        Args:
            row (int): Which row to write.
            sheet (openpyxl.worksheet.worksheet): Which sheet to write.
        """
        for col, value in enumerate(self.detail.values()):
            sheet.cell(row, col + 1, value)

# Request headers to pretend to be browsers
headers = {
    'Accept': 'application/json;text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Language': 'en-US,en;q=0.9,zh-CN;q=0.8,zh;q=0.7',
    'Connection': 'keep-alive',
    # 'Cookie': 'agentKey=6faqnsNHReKMhndB7bP4zJ; _gcl_au=1.1.713573593.1670022996; ln_or=d; _gid=GA1.2.753023053.1670022997; _pin_unauth=dWlkPU1UY3daRGcwTldNdFltRmlNQzAwTnpNMExUazFOVE10WVdKaE56azVOakUxTkRBeA; _hjSessionUser_1272670=eyJpZCI6IjJlZjA2MTY1LTAwYTgtNTgzYy1iNDI2LTgyYTUzOTY5ZGYwYSIsImNyZWF0ZWQiOjE2NzAwMjI5OTc1NjQsImV4aXN0aW5nIjp0cnVlfQ==; acceptedCookies=true; mfpLocale=en; _hjIncludedInSessionSample=0; _hjSession_1272670=eyJpZCI6IjAxNDU5OGRlLTc4MjctNDM1Yy04ZDQ0LTAxZGYyYWYwMTQ4NiIsImNyZWF0ZWQiOjE2NzAwMzMxMTQxMjUsImluU2FtcGxlIjpmYWxzZX0=; _hjAbsoluteSessionInProgress=0; _ga=GA1.2.381664351.1670022997; _derived_epik=dj0yJnU9dGh2VzYwZWlkdXJuc1dPNUxJblVZWjNVZWcwM1hFUk8mbj0zXzhoeUxCcFFwaG5JelVkdHJyVWV3Jm09MSZ0PUFBQUFBR09Lc3NrJnJtPTEmcnQ9QUFBQUFHT0tzc2smc3A9Mg; _ga_GZE6PFT72W=GS1.1.1670033109.3.1.1670034323.0.0.0',
    'If-None-Match': 'W/"H-828860823"',
    'Referer': 'https://www.artbasel.com/galleries',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Google Chrome";v="107", "Chromium";v="107", "Not=A?Brand";v="24"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Linux"',
}

# ---------- Get gallery ids -----------
# Request parameters
params = {
    # List length (0 - 100).
    'limit': 100, 
    # Start from the `offset`th item.
    'offset': 0,
    'sortBy': "sortingName",
    'statuses': 102900000
}

# Request URL to get a gallery list
url = "https://www.artbasel.com/msvc/v1/artcatalog/gallery/items"
# Array to store fetched gallery ids.
ids = []
# Whether there are more galleries
has_more = True
# While there are more galleries, keep going
while has_more:
    # Send request
    response = requests.get(url, headers=headers, params=params)
    # Parse feteched json object
    list_json = json.loads(response.text)
    # Find the gallery list
    galleries_json = list_json["items"]
    # Update has_more
    has_more = list_json["hasMore"]

    # Get each gallery's id
    for gallery_json in galleries_json:
        ids.append(gallery_json["id"])

    print(f"loading gallery {params['offset'] + 1} - {params['offset'] + len(galleries_json)}...")

    # Next page
    params["offset"] += params["limit"]

print("done")

# ------------- Get gallery details -------------
def get_detail_from_id_mp(gallery_id):
    """
    A multi-processing function to get a gallery's detail information.
    """

    # -------- Get detail --------
    gallery = Gallery(gallery_id)
    gallery.get_detail()

    # Counter
    global counter
    with counter.get_lock():
        counter.value += 1
        print(f"{counter.value} finished.")
    
    return gallery

# Initialize a counter
counter = Value('i', 0)
# Get all galleries' detail information
with Pool(20) as p:
    galleries = p.map(get_detail_from_id_mp, ids)


# ----------- Write to excel -----------------
# Initiate an excel sheet
workbook = Workbook()
sheet = workbook.active

# Sheet head
titles = ["Gallery", "Description", "Show", "Email Address", "Directors", "Adresses", "Website"]
for i in range(len(titles)):
    sheet.cell(1, i + 1, titles[i])

# Write fetched information
for index, gallery in enumerate(galleries):
    gallery.write(index + 2, sheet)

print("done")

# ------ Save the excel file --------
name = "gallery"
workbook.save(filename = name + str(int(time.time())) + ".xlsx")

